import re
import json
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import case, func
from sqlalchemy.orm import Session as DBSession

from app.models import FieldComparison, Session as SessionModel, TradeRecord
from app.models import ReportingSnapshot
from app.services.export import (
    HEADER_FILL,
    HEADER_FONT,
    PAIRING_STYLES,
    SEVERITY_FILLS,
    SEVERITY_FONTS,
    THIN_BORDER,
    UNMATCH_ROW_FILL,
    UNPAIR_ROW_FILL,
)

UNPAIR_FIELDS = {"UTI", "Other counterparty"}
DATE_IN_FILENAME_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _parse_date(date_value: Optional[str]):
    if not date_value:
        return None
    return datetime.strptime(date_value, "%Y-%m-%d").date()


def _session_business_date(session: SessionModel):
    if session.filename:
        match = DATE_IN_FILENAME_RE.search(session.filename)
        if match:
            return datetime.strptime(match.group(1), "%Y-%m-%d").date()
    created_at = session.created_at or datetime.now(timezone.utc)
    return created_at.date()


def _get_filtered_sessions(db: DBSession, date_from: Optional[str], date_to: Optional[str]) -> list[SessionModel]:
    start = _parse_date(date_from)
    end = _parse_date(date_to)
    sessions = db.query(SessionModel).order_by(SessionModel.created_at.asc()).all()
    filtered: list[SessionModel] = []
    for session in sessions:
        session_date = _session_business_date(session)
        if start and session_date < start:
            continue
        if end and session_date > end:
            continue
        filtered.append(session)
    return filtered


def _get_pairing_by_trade(db: DBSession, session_ids: list[int]) -> dict[int, dict[str, Optional[str]]]:
    if not session_ids:
        return {}

    trades = (
        db.query(TradeRecord.id, TradeRecord.has_unmatches)
        .filter(TradeRecord.session_id.in_(session_ids))
        .all()
    )
    trade_ids = [trade_id for trade_id, _ in trades]
    if not trade_ids:
        return {}

    rows = (
        db.query(FieldComparison.trade_id, FieldComparison.field_name)
        .filter(
            FieldComparison.trade_id.in_(trade_ids),
            FieldComparison.field_name.in_(UNPAIR_FIELDS),
            FieldComparison.result == "UNMATCH",
        )
        .all()
    )
    reasons_by_trade: dict[int, set[str]] = defaultdict(set)
    for trade_id, field_name in rows:
        reasons_by_trade[trade_id].add(field_name)

    result: dict[int, dict[str, Optional[str]]] = {}
    for trade_id, has_unmatches in trades:
        if reasons_by_trade.get(trade_id):
            ordered_reasons = ", ".join(sorted(reasons_by_trade[trade_id]))
            result[trade_id] = {"pairing_status": "UNPAIR", "pairing_reason": ordered_reasons}
        elif has_unmatches:
            result[trade_id] = {"pairing_status": "UNMATCH", "pairing_reason": None}
        else:
            result[trade_id] = {"pairing_status": "CLEAN", "pairing_reason": None}
    return result


def _build_overview(sessions: list[SessionModel], status_counts: Counter, pairing_counts: Counter, date_from: Optional[str], date_to: Optional[str]) -> dict:
    total_trades = sum(session.total_trades for session in sessions)
    trades_with_unmatches = sum(session.trades_with_unmatches for session in sessions)
    total_unmatches = sum(session.total_unmatches for session in sessions)
    resolved_fields = status_counts.get("RESOLVED", 0)
    return {
        "date_from": date_from,
        "date_to": date_to,
        "sessions": len(sessions),
        "total_trades": total_trades,
        "trades_with_unmatches": trades_with_unmatches,
        "total_unmatches": total_unmatches,
        "critical_count": sum(session.critical_count for session in sessions),
        "warning_count": sum(session.warning_count for session in sessions),
        "unpair_trades": pairing_counts.get("UNPAIR", 0),
        "unmatch_trades": pairing_counts.get("UNMATCH", 0),
        "clean_trades": pairing_counts.get("CLEAN", 0),
        "pending_fields": status_counts.get("PENDING", 0),
        "resolved_fields": status_counts.get("RESOLVED", 0),
        "in_negotiation_fields": status_counts.get("IN_NEGOTIATION", 0),
        "excluded_fields": status_counts.get("EXCLUDED", 0),
        "quality_rate": round(((total_trades - trades_with_unmatches) / total_trades) * 100, 2) if total_trades else 0,
        "resolution_rate": round((resolved_fields / total_unmatches) * 100, 2) if total_unmatches else 0,
    }


def _build_comparison_to_previous_period(db: DBSession, current_overview: dict, date_from: Optional[str], date_to: Optional[str]) -> dict | None:
    start = _parse_date(date_from)
    end = _parse_date(date_to)
    if not start or not end or end < start:
      return None

    period_days = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=period_days - 1)
    previous_sessions = _get_filtered_sessions(db, previous_start.isoformat(), previous_end.isoformat())
    previous_session_ids = [session.id for session in previous_sessions]
    previous_status_counts = Counter(
        dict(
            db.query(FieldComparison.status, func.count(FieldComparison.id))
            .filter(FieldComparison.session_id.in_(previous_session_ids))
            .group_by(FieldComparison.status)
            .all()
        )
    ) if previous_session_ids else Counter()
    previous_pairing_counts = Counter(
        item["pairing_status"] for item in _get_pairing_by_trade(db, previous_session_ids).values()
    ) if previous_session_ids else Counter()
    previous_overview = _build_overview(
        previous_sessions,
        previous_status_counts,
        previous_pairing_counts,
        previous_start.isoformat(),
        previous_end.isoformat(),
    )

    def _delta(metric: str) -> dict:
        current_value = float(current_overview.get(metric, 0) or 0)
        previous_value = float(previous_overview.get(metric, 0) or 0)
        abs_delta = round(current_value - previous_value, 2)
        pct_delta = round(((current_value - previous_value) / previous_value) * 100, 2) if previous_value else None
        return {
            "previous": previous_value,
            "current": current_value,
            "abs": abs_delta,
            "pct": pct_delta,
        }

    return {
        "previous_date_from": previous_start.isoformat(),
        "previous_date_to": previous_end.isoformat(),
        "deltas": {
            "total_unmatches": _delta("total_unmatches"),
            "critical_count": _delta("critical_count"),
            "unpair_trades": _delta("unpair_trades"),
            "quality_rate": _delta("quality_rate"),
            "resolution_rate": _delta("resolution_rate"),
            "pending_fields": _delta("pending_fields"),
        },
    }


def _build_risk_residual(open_items: list[dict], critical_open_items: list[dict], trade_summaries: list[dict], top_fields: list[dict]) -> dict:
    unresolved_unpair_trades = sum(
        1 for trade in trade_summaries
        if trade.get("pairing_status") == "UNPAIR" and trade.get("has_unmatches")
    )
    concentration_top_field_pct = round((top_fields[0]["count"] / len(open_items)) * 100, 2) if open_items and top_fields else 0

    if critical_open_items or unresolved_unpair_trades >= 25:
        level = "ALTO"
    elif len(open_items) >= 100:
        level = "MEDIO"
    else:
        level = "BAJO"

    return {
        "level": level,
        "open_items": len(open_items),
        "critical_open_items": len(critical_open_items),
        "unresolved_unpair_trades": unresolved_unpair_trades,
        "top_field_concentration_pct": concentration_top_field_pct,
        "summary": (
            f"Riesgo residual {level.lower()} con {len(critical_open_items)} incidencias críticas abiertas, "
            f"{unresolved_unpair_trades} operaciones UNPAIR sin resolver y una concentración del {concentration_top_field_pct}% "
            "en el principal campo con incidencia."
        ),
    }


def build_regulatory_report_preview(db: DBSession, date_from: Optional[str], date_to: Optional[str]) -> dict:
    sessions = _get_filtered_sessions(db, date_from, date_to)
    session_ids = [session.id for session in sessions]
    generated_at = datetime.now(timezone.utc)

    if not session_ids:
        return {
            "date_from": date_from,
            "date_to": date_to,
            "generated_at": generated_at,
            "sessions": 0,
            "filenames": [],
            "overview": _build_overview([], Counter(), Counter(), date_from, date_to),
            "daily_summary": [],
            "top_fields": [],
            "top_counterparties": [],
            "open_items_count": 0,
            "critical_open_items_count": 0,
            "open_items": [],
            "comparison_to_previous_period": None,
            "risk_residual": {
                "level": "BAJO",
                "open_items": 0,
                "critical_open_items": 0,
                "unresolved_unpair_trades": 0,
                "top_field_concentration_pct": 0,
                "summary": "No hay backlog abierto en el periodo seleccionado.",
            },
            "trade_summaries": [],
            "field_details": [],
        }

    trade_rows = (
        db.query(TradeRecord, SessionModel)
        .join(SessionModel, SessionModel.id == TradeRecord.session_id)
        .filter(TradeRecord.session_id.in_(session_ids))
        .order_by(SessionModel.created_at.asc(), TradeRecord.row_number.asc())
        .all()
    )
    trade_pairing = _get_pairing_by_trade(db, session_ids)
    trade_ids = [trade.id for trade, _session in trade_rows]

    status_counts = Counter(
        dict(
            db.query(FieldComparison.status, func.count(FieldComparison.id))
            .filter(FieldComparison.session_id.in_(session_ids))
            .group_by(FieldComparison.status)
            .all()
        )
    )
    pairing_counts = Counter(item["pairing_status"] for item in trade_pairing.values())
    overview = _build_overview(sessions, status_counts, pairing_counts, date_from, date_to)

    daily_buckets: dict[str, dict] = {}
    for session in sessions:
        day = _session_business_date(session).isoformat()
        bucket = daily_buckets.setdefault(day, {
            "date": day,
            "sessions": 0,
            "total_trades": 0,
            "trades_with_unmatches": 0,
            "unpair_trades": 0,
            "total_unmatches": 0,
            "critical_count": 0,
            "warning_count": 0,
            "resolved_fields": 0,
            "pending_fields": 0,
        })
        bucket["sessions"] += 1
        bucket["total_trades"] += session.total_trades
        bucket["trades_with_unmatches"] += session.trades_with_unmatches
        bucket["total_unmatches"] += session.total_unmatches
        bucket["critical_count"] += session.critical_count
        bucket["warning_count"] += session.warning_count

    if trade_ids:
        daily_trade_counts: dict[str, Counter] = defaultdict(Counter)
        for trade, session in trade_rows:
            day = _session_business_date(session).isoformat()
            daily_trade_counts[day][trade_pairing.get(trade.id, {}).get("pairing_status") or "CLEAN"] += 1

        daily_status_rows = (
            db.query(FieldComparison.status, TradeRecord.session_id, func.count(FieldComparison.id))
            .join(TradeRecord, TradeRecord.id == FieldComparison.trade_id)
            .filter(FieldComparison.trade_id.in_(trade_ids))
            .group_by(FieldComparison.status, TradeRecord.session_id)
            .all()
        )
        for status, session_id, count in daily_status_rows:
            session = next((item for item in sessions if item.id == session_id), None)
            if not session:
                continue
            day = _session_business_date(session).isoformat()
            if status == "RESOLVED":
                daily_buckets[day]["resolved_fields"] += count
            if status == "PENDING":
                daily_buckets[day]["pending_fields"] += count

        for day, counts in daily_trade_counts.items():
            daily_buckets[day]["unpair_trades"] = counts.get("UNPAIR", 0)

    top_field_rows = (
        db.query(
            FieldComparison.field_name,
            FieldComparison.table_number,
            func.count(FieldComparison.id).label("count"),
            func.sum(case((FieldComparison.severity == "CRITICAL", 1), else_=0)).label("critical_count"),
            func.sum(case((FieldComparison.severity == "WARNING", 1), else_=0)).label("warning_count"),
        )
        .filter(
            FieldComparison.session_id.in_(session_ids),
            FieldComparison.result == "UNMATCH",
        )
        .group_by(FieldComparison.field_name, FieldComparison.table_number)
        .order_by(func.count(FieldComparison.id).desc())
        .limit(10)
        .all()
    )
    top_fields = [
        {
            "field_name": field_name,
            "table_number": table_number,
            "count": count,
            "critical_count": critical_count or 0,
            "warning_count": warning_count or 0,
        }
        for field_name, table_number, count, critical_count, warning_count in top_field_rows
    ]

    counterparty_buckets: dict[tuple[str, str], dict] = {}
    for session in sessions:
        key = (session.emisor_name or "—", session.receptor_name or "—")
        bucket = counterparty_buckets.setdefault(key, {
            "emisor_name": key[0],
            "receptor_name": key[1],
            "sessions": 0,
            "total_trades": 0,
            "total_unmatches": 0,
            "critical_count": 0,
        })
        bucket["sessions"] += 1
        bucket["total_trades"] += session.total_trades
        bucket["total_unmatches"] += session.total_unmatches
        bucket["critical_count"] += session.critical_count
    top_counterparties = sorted(counterparty_buckets.values(), key=lambda item: item["total_unmatches"], reverse=True)[:10]

    trade_summaries = []
    trade_lookup: dict[int, dict] = {}
    for trade, session in trade_rows:
        business_date = _session_business_date(session).isoformat()
        pairing = trade_pairing.get(trade.id, {"pairing_status": "CLEAN", "pairing_reason": None})
        item = {
            "business_date": business_date,
            "session_id": trade.session_id,
            "trade_id": trade.id,
            "row_number": trade.row_number,
            "uti": trade.uti,
            "sft_type": trade.sft_type,
            "action_type": trade.action_type,
            "emisor_name": session.emisor_name,
            "receptor_name": session.receptor_name,
            "pairing_status": pairing["pairing_status"],
            "pairing_reason": pairing["pairing_reason"],
            "total_fields": trade.total_fields,
            "total_unmatches": trade.total_unmatches,
            "critical_count": trade.critical_count,
            "warning_count": trade.warning_count,
            "has_unmatches": trade.has_unmatches,
        }
        trade_summaries.append(item)
        trade_lookup[trade.id] = item

    field_rows = (
        db.query(FieldComparison)
        .filter(
            FieldComparison.session_id.in_(session_ids),
            FieldComparison.result == "UNMATCH",
        )
        .order_by(FieldComparison.session_id.asc(), FieldComparison.trade_id.asc(), FieldComparison.table_number.asc(), FieldComparison.field_number.asc())
        .all()
    )
    field_details = []
    open_items = []
    for field in field_rows:
        trade_meta = trade_lookup.get(field.trade_id, {})
        item = {
            "business_date": trade_meta.get("business_date"),
            "session_id": field.session_id,
            "trade_id": field.trade_id,
            "row_number": trade_meta.get("row_number"),
            "uti": trade_meta.get("uti"),
            "sft_type": trade_meta.get("sft_type"),
            "action_type": trade_meta.get("action_type"),
            "pairing_status": trade_meta.get("pairing_status", "CLEAN"),
            "pairing_reason": trade_meta.get("pairing_reason"),
            "field_name": field.field_name,
            "table_number": field.table_number,
            "field_number": field.field_number,
            "obligation": field.obligation,
            "emisor_value": field.emisor_value,
            "receptor_value": field.receptor_value,
            "result": field.result,
            "severity": field.severity,
            "root_cause": field.root_cause,
            "status": field.status,
            "assignee": field.assignee,
            "notes": field.notes,
            "validated": field.validated,
            "updated_at": field.updated_at,
        }
        field_details.append(item)
        if field.status not in {"RESOLVED", "EXCLUDED"}:
            open_items.append(item)

    critical_open_items = [item for item in open_items if item["severity"] == "CRITICAL"]
    comparison_to_previous_period = _build_comparison_to_previous_period(db, overview, date_from, date_to)
    risk_residual = _build_risk_residual(open_items, critical_open_items, trade_summaries, top_fields)

    return {
        "date_from": date_from,
        "date_to": date_to,
        "generated_at": generated_at,
        "sessions": len(sessions),
        "filenames": [session.filename for session in sessions if session.filename],
        "overview": overview,
        "daily_summary": sorted(daily_buckets.values(), key=lambda item: item["date"]),
        "top_fields": top_fields,
        "top_counterparties": top_counterparties,
        "open_items_count": len(open_items),
        "critical_open_items_count": len(critical_open_items),
        "open_items": open_items[:50],
        "comparison_to_previous_period": comparison_to_previous_period,
        "risk_residual": risk_residual,
        "trade_summaries": trade_summaries,
        "field_details": field_details,
    }


def build_regulatory_narrative_fallback(report: dict) -> str:
    overview = report["overview"]
    top_fields = report.get("top_fields", [])[:5]
    top_counterparties = report.get("top_counterparties", [])[:5]
    open_items = report.get("open_items", [])[:5]
    risk_residual = report.get("risk_residual") or {}
    previous_comparison = report.get("comparison_to_previous_period") or {}
    deltas = previous_comparison.get("deltas", {})
    top_fields_text = "\n".join(
        f"- {item['field_name']} (Tabla {item['table_number']}): {item['count']} discrepancias"
        for item in top_fields
    ) or "- No hay campos destacados"
    counterparties_text = "\n".join(
        f"- {item['emisor_name']} vs {item['receptor_name']}: {item['total_unmatches']} discrepancias"
        for item in top_counterparties
    ) or "- No hay contrapartes destacadas"
    open_items_text = "\n".join(
        f"- {item['business_date']} · UTI {item.get('uti') or '—'} · {item['field_name']} · {item['severity']} · {item['status']}"
        for item in open_items
    ) or "- No hay incidencias abiertas en la muestra"
    comparison_text = ""
    if previous_comparison:
        comparison_text = (
            "\n## Comparación con el periodo anterior\n"
            f"- Periodo anterior: {previous_comparison.get('previous_date_from')} a {previous_comparison.get('previous_date_to')}\n"
            f"- Delta discrepancias: {deltas.get('total_unmatches', {}).get('abs', 0):+}\n"
            f"- Delta críticas: {deltas.get('critical_count', {}).get('abs', 0):+}\n"
            f"- Delta UNPAIR: {deltas.get('unpair_trades', {}).get('abs', 0):+}\n"
            f"- Delta calidad: {deltas.get('quality_rate', {}).get('abs', 0):+}\n"
            f"- Delta resolución: {deltas.get('resolution_rate', {}).get('abs', 0):+}\n"
        )

    return (
        "# Informe regulatorio SFTR\n\n"
        f"**Periodo:** {report.get('date_from') or 'inicio disponible'} a {report.get('date_to') or 'fin disponible'}\n"
        f"**Sesiones incluidas:** {report.get('sessions', 0)}\n"
        f"**Operaciones procesadas:** {overview.get('total_trades', 0)}\n"
        f"**Operaciones con discrepancias:** {overview.get('trades_with_unmatches', 0)}\n"
        f"**UNPAIR:** {overview.get('unpair_trades', 0)}\n"
        f"**UNMATCH:** {overview.get('unmatch_trades', 0)}\n"
        f"**Críticas:** {overview.get('critical_count', 0)}\n"
        f"**Pendientes:** {overview.get('pending_fields', 0)}\n"
        f"**Resueltas:** {overview.get('resolved_fields', 0)}\n"
        f"**Calidad:** {overview.get('quality_rate', 0)}%\n"
        f"**Resolución:** {overview.get('resolution_rate', 0)}%\n\n"
        "## Lectura ejecutiva\n"
        f"- El periodo presenta {overview.get('total_unmatches', 0)} discrepancias y {report.get('critical_open_items_count', 0)} incidencias críticas abiertas.\n"
        f"- La carga operativa se concentra en {overview.get('unpair_trades', 0)} operaciones `UNPAIR` y {report.get('open_items_count', 0)} elementos abiertos.\n"
        "- El anexo estructurado adjunto contiene el detalle auditable de operaciones, backlog y evolución diaria.\n\n"
        "## Campos con más incidencia\n"
        f"{top_fields_text}\n\n"
        "## Contrapartes más expuestas\n"
        f"{counterparties_text}\n\n"
        "## Open items destacados\n"
        f"{open_items_text}\n\n"
        f"{comparison_text}\n"
        "## Riesgo residual\n"
        f"- {risk_residual.get('summary', 'Sin evaluación de riesgo residual disponible.')}\n\n"
        "## Recomendación operativa\n"
        "- Priorizar los `UNPAIR` y las incidencias críticas abiertas, consolidar responsables y reprocesar las sesiones afectadas tras la remediación."
    )


def snapshot_payload_to_preview(payload: dict) -> dict:
    preview = dict(payload)
    preview.setdefault("generated_at", datetime.now(timezone.utc).isoformat())
    return preview


def serialize_report_for_snapshot(report: dict) -> str:
    serializable = dict(report)
    serializable["generated_at"] = (
        report["generated_at"].isoformat()
        if isinstance(report.get("generated_at"), datetime)
        else report.get("generated_at")
    )
    for item in serializable.get("open_items", []):
        if isinstance(item.get("updated_at"), datetime):
            item["updated_at"] = item["updated_at"].isoformat()
    for item in serializable.get("field_details", []):
        if isinstance(item.get("updated_at"), datetime):
            item["updated_at"] = item["updated_at"].isoformat()
    return json.dumps(serializable)


def deserialize_snapshot_payload(snapshot: ReportingSnapshot) -> dict:
    return json.loads(snapshot.payload_json)


def generate_regulatory_xlsx(report: dict) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    _populate_executive_summary(ws_summary, report)

    ws_trades = wb.create_sheet("Trades Summary")
    _populate_trades_sheet(ws_trades, report["trade_summaries"])

    open_items = [item for item in report["field_details"] if item["status"] not in {"RESOLVED", "EXCLUDED"}]
    ws_open = wb.create_sheet("Open Items")
    _populate_field_sheet(ws_open, "Open Items", open_items)

    critical_open_items = [item for item in open_items if item["severity"] == "CRITICAL"]
    ws_critical = wb.create_sheet("Critical Open Items")
    _populate_field_sheet(ws_critical, "Critical Open Items", critical_open_items)

    ws_field_detail = wb.create_sheet("Field Detail")
    _populate_field_sheet(ws_field_detail, "Field Detail", report["field_details"])

    ws_daily = wb.create_sheet("Daily Trend")
    _populate_daily_sheet(ws_daily, report["daily_summary"])

    ws_fields = wb.create_sheet("Top Fields")
    _populate_top_fields_sheet(ws_fields, report["top_fields"])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _populate_executive_summary(ws, report: dict) -> None:
    ws["A1"] = "Regulatory Reporting Summary"
    ws["A1"].font = Font(bold=True, size=14, color="2B3544")

    overview = report["overview"]
    rows = [
        ("Date from", report.get("date_from") or "—"),
        ("Date to", report.get("date_to") or "—"),
        ("Generated at", str(report["generated_at"])),
        ("Sessions", report["sessions"]),
        ("Total trades", overview["total_trades"]),
        ("Trades with discrepancies", overview["trades_with_unmatches"]),
        ("UNPAIR", overview["unpair_trades"]),
        ("UNMATCH", overview["unmatch_trades"]),
        ("Clean trades", overview["clean_trades"]),
        ("Total discrepancies", overview["total_unmatches"]),
        ("Critical", overview["critical_count"]),
        ("Warning", overview["warning_count"]),
        ("Pending fields", overview["pending_fields"]),
        ("Resolved fields", overview["resolved_fields"]),
        ("Quality rate %", overview["quality_rate"]),
        ("Resolution rate %", overview["resolution_rate"]),
        ("Open items", report["open_items_count"]),
        ("Critical open items", report["critical_open_items_count"]),
    ]
    for idx, (label, value) in enumerate(rows, start=3):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)

    ws["D3"] = "Included files"
    ws["D3"].font = Font(bold=True, size=12)
    for idx, filename in enumerate(report["filenames"], start=4):
        ws.cell(row=idx, column=4, value=filename)

    ws["F3"] = "Top counterparties"
    ws["F3"].font = Font(bold=True, size=12)
    for idx, item in enumerate(report["top_counterparties"][:8], start=4):
        ws.cell(row=idx, column=6, value=f'{item["emisor_name"]} / {item["receptor_name"]}').font = Font(bold=True)
        ws.cell(row=idx, column=7, value=item["total_unmatches"])

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 16


def _populate_trades_sheet(ws, trades: list[dict]) -> None:
    headers = [
        "Business Date", "Session ID", "Trade ID", "Row #", "UTI", "SFT Type", "Action Type",
        "Emisor", "Receptor", "Pairing Status", "Pairing Reason", "Total Fields", "Total Unmatches",
        "Critical", "Warning", "Has Unmatches",
    ]
    _write_headers(ws, headers)
    for row_index, trade in enumerate(trades, start=2):
        values = [
            trade["business_date"], trade["session_id"], trade["trade_id"], trade["row_number"], trade["uti"],
            trade["sft_type"], trade["action_type"], trade["emisor_name"], trade["receptor_name"],
            trade["pairing_status"], trade["pairing_reason"], trade["total_fields"], trade["total_unmatches"],
            trade["critical_count"], trade["warning_count"], "Yes" if trade["has_unmatches"] else "No",
        ]
        _write_trade_row(ws, row_index, values, trade["pairing_status"])
    _set_widths(ws, [14, 10, 10, 8, 20, 12, 12, 18, 18, 14, 28, 12, 14, 10, 10, 12])


def _populate_field_sheet(ws, title: str, fields: list[dict]) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="2B3544")
    headers = [
        "Business Date", "Session ID", "Trade ID", "Row #", "UTI", "SFT Type", "Action Type", "Pairing Status",
        "Pairing Reason", "Table", "Field #", "Field Name", "Obligation", "Emisor Value", "Receptor Value",
        "Result", "Severity", "Root Cause", "Status", "Assignee", "Notes", "Validated", "Updated At",
    ]
    _write_headers(ws, headers, row=3)
    for row_index, field in enumerate(fields, start=4):
        values = [
            field["business_date"], field["session_id"], field["trade_id"], field["row_number"], field["uti"],
            field["sft_type"], field["action_type"], field["pairing_status"], field["pairing_reason"],
            field["table_number"], field["field_number"], field["field_name"], field["obligation"],
            field["emisor_value"], field["receptor_value"], field["result"], field["severity"], field["root_cause"],
            field["status"], field["assignee"], field["notes"], "Yes" if field["validated"] else "No",
            str(field["updated_at"] or ""),
        ]
        _write_field_row(ws, row_index, values, field["pairing_status"], field["severity"])
    _set_widths(ws, [14, 10, 10, 8, 20, 12, 12, 14, 28, 8, 8, 34, 10, 20, 20, 12, 12, 20, 16, 18, 32, 10, 20])


def _populate_daily_sheet(ws, rows: list[dict]) -> None:
    headers = [
        "Date", "Sessions", "Total Trades", "Trades with Discrepancies", "UNPAIR", "Total Discrepancies",
        "Critical", "Warning", "Resolved Fields", "Pending Fields",
    ]
    _write_headers(ws, headers)
    for row_index, item in enumerate(rows, start=2):
        values = [
            item["date"], item["sessions"], item["total_trades"], item["trades_with_unmatches"], item["unpair_trades"],
            item["total_unmatches"], item["critical_count"], item["warning_count"], item["resolved_fields"], item["pending_fields"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_index > 1 else "left")
    _set_widths(ws, [14, 10, 12, 22, 10, 18, 10, 10, 16, 16])


def _populate_top_fields_sheet(ws, rows: list[dict]) -> None:
    headers = ["Field Name", "Table", "Count", "Critical Count", "Warning Count"]
    _write_headers(ws, headers)
    for row_index, item in enumerate(rows, start=2):
        values = [item["field_name"], item["table_number"], item["count"], item["critical_count"], item["warning_count"]]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_index > 1 else "left")
    _set_widths(ws, [36, 10, 10, 14, 14])


def _write_headers(ws, headers: list[str], row: int = 1) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _write_trade_row(ws, row_index: int, values: list, pairing_status: str) -> None:
    row_fill = UNPAIR_ROW_FILL if pairing_status == "UNPAIR" else UNMATCH_ROW_FILL if pairing_status == "UNMATCH" else None
    pairing_fill_color, pairing_font_color = PAIRING_STYLES.get(pairing_status, PAIRING_STYLES[""])
    for col_index, value in enumerate(values, start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill
        if col_index == 10:
            cell.fill = PatternFill(start_color=pairing_fill_color, end_color=pairing_fill_color, fill_type="solid")
            cell.font = Font(color=pairing_font_color, bold=bool(pairing_status))
        cell.alignment = Alignment(horizontal="center" if col_index in (2, 3, 4, 10, 12, 13, 14, 15, 16) else "left")


def _write_field_row(ws, row_index: int, values: list, pairing_status: str, severity: str) -> None:
    row_fill = UNPAIR_ROW_FILL if pairing_status == "UNPAIR" else UNMATCH_ROW_FILL if pairing_status == "UNMATCH" else None
    pairing_fill_color, pairing_font_color = PAIRING_STYLES.get(pairing_status, PAIRING_STYLES[""])
    for col_index, value in enumerate(values, start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill
        if col_index == 8:
            cell.fill = PatternFill(start_color=pairing_fill_color, end_color=pairing_fill_color, fill_type="solid")
            cell.font = Font(color=pairing_font_color, bold=bool(pairing_status))
        if col_index == 17:
            cell.fill = SEVERITY_FILLS.get(severity, SEVERITY_FILLS["NONE"])
            cell.font = SEVERITY_FONTS.get(severity, SEVERITY_FONTS["NONE"])
        cell.alignment = Alignment(horizontal="center" if col_index in (2, 3, 4, 8, 10, 11, 16, 17, 22) else "left")


def _set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[_excel_column_name(idx)].width = width


def _excel_column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name
