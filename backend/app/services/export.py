from collections import Counter
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SEVERITY_FILLS = {
    "CRITICAL": PatternFill(start_color="FCEBEB", end_color="FCEBEB", fill_type="solid"),
    "WARNING": PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid"),
    "INFO": PatternFill(start_color="E6F1FB", end_color="E6F1FB", fill_type="solid"),
    "NONE": PatternFill(start_color="EAF3DE", end_color="EAF3DE", fill_type="solid"),
}

SEVERITY_FONTS = {
    "CRITICAL": Font(color="A32D2D", bold=True),
    "WARNING": Font(color="854F0B", bold=True),
    "INFO": Font(color="185FA5"),
    "NONE": Font(color="3B6D11"),
}

PAIRING_STYLES = {
    "UNPAIR": ("EDE9FE", "6D28D9"),
    "UNMATCH": ("FEE2E2", "B91C1C"),
    "": ("F4F4F5", "52525B"),
}
UNPAIR_ROW_FILL = PatternFill(start_color="F5EFFF", end_color="F5EFFF", fill_type="solid")
UNMATCH_ROW_FILL = PatternFill(start_color="FFF4F4", end_color="FFF4F4", fill_type="solid")

HEADER_FILL = PatternFill(start_color="2B3544", end_color="2B3544", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

DETAIL_HEADERS = [
    "Trade ID",
    "Row #",
    "UTI",
    "SFT Type",
    "Action Type",
    "Pairing Status",
    "Pairing Reason",
    "Table",
    "Field #",
    "Field Name",
    "Obligation",
    "Emisor Value",
    "Receptor Value",
    "Result",
    "Severity",
    "Root Cause",
    "Status",
    "Assignee",
    "Notes",
    "Validated",
]

TRADE_SUMMARY_HEADERS = [
    "Trade ID",
    "Row #",
    "UTI",
    "SFT Type",
    "Action Type",
    "Pairing Status",
    "Pairing Reason",
    "Total Fields",
    "Total Unmatches",
    "Critical",
    "Warning",
    "Emisor LEI",
    "Receptor LEI",
]


def generate_xlsx(session_data: dict, field_results: list, only_unmatches: bool = False) -> bytes:
    wb = Workbook()
    all_results = [
        item for item in field_results
        if not only_unmatches or item.get("result") == "UNMATCH"
    ]

    ws_all = wb.active
    ws_all.title = "All Results"
    _populate_detail_sheet(ws_all, "SFTR Comparison Export", session_data, all_results)

    unpair_rows = [item for item in all_results if item.get("pairing_status") == "UNPAIR"]
    unmatch_rows = [item for item in all_results if item.get("pairing_status") == "UNMATCH"]

    ws_unpair = wb.create_sheet("Unpair")
    _populate_detail_sheet(ws_unpair, "UNPAIR Operations", session_data, unpair_rows)

    ws_unmatch = wb.create_sheet("Unmatch")
    _populate_detail_sheet(ws_unmatch, "UNMATCH Operations", session_data, unmatch_rows)

    ws_trades = wb.create_sheet("Trades Summary")
    _populate_trade_summary_sheet(ws_trades, session_data.get("trade_summaries", []))

    ws_summary = wb.create_sheet("Summary")
    _populate_summary_sheet(
        ws_summary,
        session_data,
        all_results,
        unpair_rows,
        unmatch_rows,
        session_data.get("trade_summaries", []),
    )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _populate_detail_sheet(ws, title: str, session_data: dict, rows: list[dict]) -> None:
    ws.merge_cells("A1:U1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="2B3544")
    title_cell.alignment = Alignment(horizontal="left")

    ws["A2"] = "Session ID:"
    ws["B2"] = session_data.get("id", "")
    ws["A3"] = "Session Date:"
    ws["B3"] = session_data.get("created_at", "")
    ws["A4"] = "SFT Type:"
    ws["B4"] = session_data.get("sft_type", "")
    ws["C4"] = "Action Type:"
    ws["D4"] = session_data.get("action_type", "")
    ws["A5"] = "Emisor:"
    ws["B5"] = session_data.get("emisor_name", "")
    ws["C5"] = "Receptor:"
    ws["D5"] = session_data.get("receptor_name", "")
    ws["A6"] = "Filename:"
    ws["B6"] = session_data.get("filename", "")

    for row_index in range(2, 7):
        ws.cell(row=row_index, column=1).font = Font(bold=True, size=10)
        ws.cell(row=row_index, column=3).font = Font(bold=True, size=10)

    header_row = 8
    for col_index, header in enumerate(DETAIL_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    current_row = header_row + 1
    for item in rows:
        severity = item.get("severity", "NONE")
        severity_fill = SEVERITY_FILLS.get(severity, SEVERITY_FILLS["NONE"])
        severity_font = SEVERITY_FONTS.get(severity, SEVERITY_FONTS["NONE"])
        pairing_status = item.get("pairing_status") or ""
        pairing_fill_color, pairing_font_color = PAIRING_STYLES.get(pairing_status, PAIRING_STYLES[""])
        row_fill = UNPAIR_ROW_FILL if pairing_status == "UNPAIR" else UNMATCH_ROW_FILL if pairing_status == "UNMATCH" else None

        values = [
            item.get("trade_id"),
            item.get("row_number"),
            item.get("uti", ""),
            item.get("sft_type", ""),
            item.get("action_type", ""),
            pairing_status,
            item.get("pairing_reason", ""),
            item.get("table_number"),
            item.get("field_number"),
            item.get("field_name"),
            item.get("obligation"),
            item.get("emisor_value", ""),
            item.get("receptor_value", ""),
            item.get("result"),
            severity,
            item.get("root_cause", ""),
            item.get("status", ""),
            item.get("assignee", ""),
            item.get("notes", ""),
            "Yes" if item.get("validated") else "No",
        ]

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_index, value=value)
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill
            if col_index == 6:
                cell.fill = PatternFill(start_color=pairing_fill_color, end_color=pairing_fill_color, fill_type="solid")
                cell.font = Font(color=pairing_font_color, bold=bool(pairing_status))
            elif col_index in (14, 15):
                cell.fill = severity_fill
                cell.font = severity_font
            cell.alignment = Alignment(horizontal="center" if col_index in (1, 2, 6, 8, 9, 11, 14, 15, 20) else "left")
        current_row += 1

    widths = [10, 8, 20, 12, 12, 14, 28, 8, 8, 36, 10, 24, 24, 12, 12, 22, 16, 18, 28, 10]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[_excel_column_name(idx)].width = width


def _populate_trade_summary_sheet(ws, trade_summaries: list[dict]) -> None:
    ws["A1"] = "Trades Summary"
    ws["A1"].font = Font(bold=True, size=14, color="2B3544")

    header_row = 3
    for col_index, header in enumerate(TRADE_SUMMARY_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    current_row = header_row + 1
    for trade in trade_summaries:
        pairing_status = trade.get("pairing_status") or ""
        pairing_fill_color, pairing_font_color = PAIRING_STYLES.get(pairing_status, PAIRING_STYLES[""])
        row_fill = UNPAIR_ROW_FILL if pairing_status == "UNPAIR" else UNMATCH_ROW_FILL if pairing_status == "UNMATCH" else None

        values = [
            trade.get("trade_id"),
            trade.get("row_number"),
            trade.get("uti"),
            trade.get("sft_type"),
            trade.get("action_type"),
            pairing_status,
            trade.get("pairing_reason", ""),
            trade.get("total_fields"),
            trade.get("total_unmatches"),
            trade.get("critical_count"),
            trade.get("warning_count"),
            trade.get("emisor_lei", ""),
            trade.get("receptor_lei", ""),
        ]

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_index, value=value)
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill
            if col_index == 6:
                cell.fill = PatternFill(start_color=pairing_fill_color, end_color=pairing_fill_color, fill_type="solid")
                cell.font = Font(color=pairing_font_color, bold=bool(pairing_status))
            cell.alignment = Alignment(horizontal="center" if col_index in (1, 2, 6, 8, 9, 10, 11) else "left")
        current_row += 1

    widths = [10, 8, 18, 12, 12, 14, 28, 12, 14, 10, 10, 22, 22]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[_excel_column_name(idx)].width = width


def _populate_summary_sheet(
    ws,
    session_data: dict,
    all_rows: list[dict],
    unpair_rows: list[dict],
    unmatch_rows: list[dict],
    trade_summaries: list[dict],
) -> None:
    ws["A1"] = "SFTR Export Summary"
    ws["A1"].font = Font(bold=True, size=14, color="2B3544")

    severity_counts = Counter(row.get("severity", "NONE") for row in all_rows if row.get("result") == "UNMATCH")
    pairing_counts = Counter((trade.get("pairing_status") or "CLEAN") for trade in trade_summaries)

    summary_items = [
        ("Session ID", session_data.get("id", "")),
        ("Filename", session_data.get("filename", "")),
        ("Session Date", session_data.get("created_at", "")),
        ("SFT Type", session_data.get("sft_type", "")),
        ("Action Type", session_data.get("action_type", "")),
        ("Emisor", session_data.get("emisor_name", "")),
        ("Receptor", session_data.get("receptor_name", "")),
        ("Total trades", session_data.get("total_trades", 0)),
        ("Total unmatches", session_data.get("total_unmatches", 0)),
        ("Critical count", session_data.get("critical_count", 0)),
        ("Rows in All Results", len(all_rows)),
        ("Rows in Unpair", len(unpair_rows)),
        ("Rows in Unmatch", len(unmatch_rows)),
        ("Trades marked Unpair", pairing_counts.get("UNPAIR", 0)),
        ("Trades marked Unmatch", pairing_counts.get("UNMATCH", 0)),
        ("Trades clean", pairing_counts.get("CLEAN", 0)),
    ]

    for idx, (label, value) in enumerate(summary_items, start=3):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)

    ws["D3"] = "Legend"
    ws["D3"].font = Font(bold=True, size=12)
    legend = [
        ("UNPAIR", "Operations where UTI and/or Other counterparty do not pair"),
        ("UNMATCH", "Operations with discrepancies outside the pairing fields"),
    ]
    for idx, (label, text) in enumerate(legend, start=4):
        ws.cell(row=idx, column=4, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=5, value=text)

    ws["D8"] = "Unmatch Severity Breakdown"
    ws["D8"].font = Font(bold=True, size=12)
    severity_rows = [
        ("CRITICAL", severity_counts.get("CRITICAL", 0)),
        ("WARNING", severity_counts.get("WARNING", 0)),
        ("INFO", severity_counts.get("INFO", 0)),
    ]
    for idx, (label, value) in enumerate(severity_rows, start=9):
        ws.cell(row=idx, column=4, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=5, value=value)

    ws["D14"] = "Trade Pairing Breakdown"
    ws["D14"].font = Font(bold=True, size=12)
    trade_rows = [
        ("UNPAIR", pairing_counts.get("UNPAIR", 0)),
        ("UNMATCH", pairing_counts.get("UNMATCH", 0)),
        ("CLEAN", pairing_counts.get("CLEAN", 0)),
    ]
    for idx, (label, value) in enumerate(trade_rows, start=15):
        ws.cell(row=idx, column=4, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=5, value=value)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 56


def _excel_column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name
